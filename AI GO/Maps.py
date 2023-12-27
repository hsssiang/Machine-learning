import codecs
f = codecs.open("test_data.txt", 'r', 'utf-8') #記得以utf-8開啟文字檔
location = f.read()
f.close()

places = tuple(location.split('\n'))
URL = []
for i in places:
    URL.append("https://www.google.com/maps/place?q=" + i)\

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'longitude'
ws['B1'] = 'latitude'

def STR_to_NUM(data):
    line = tuple(data.split(',')) #註1
    num1 = float(line[1])
    num2 = float(line[2])
    line = [num1, num2]
    return line

def coordination(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    text = soup.prettify() #text 包含了html的內容
    initial_pos = text.find(";window.APP_INITIALIZATION_STATE")
    #尋找;window.APP_INITIALIZATION_STATE所在位置
    data = text[initial_pos+36:initial_pos+85] #將其後的參數進行存取
    num_data = STR_to_NUM(data)
    ws.append(num_data) #將經緯度存到Excel裡

import requests
from bs4 import BeautifulSoup
for i in URL:
    coordination(i)
    wb.save('test.xlsx') #自行決定檔名